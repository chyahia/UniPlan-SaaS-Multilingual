import os
import json
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from io import BytesIO
from flask_babel import _

# استدعاء دوال إدارة البيانات السحابية (الجديدة)
from app.resit_data_manager import (
    load_full_db, save_full_db,
    add_teacher, remove_teacher, edit_teacher_name,
    add_room, remove_room, edit_room_name,
    add_level, remove_level, edit_level_name,
    add_subject, remove_subject, edit_subject_name,
    update_complex_state
)

# ملاحظة: سنقوم بربط الخوارزمية لاحقاً بـ Celery
# from app.solver import run_distribution
# from app.exporter import generate_levels_word, generate_teachers_word

# إنشاء الـ Blueprint الخاص بالامتحانات الاستدراكية
resit_exams_bp = Blueprint('resit_exams', __name__, url_prefix='/resit_exams')

# ==========================================
# 🌟 حماية المسارات (التأكد من تسجيل الدخول)
# ==========================================
@resit_exams_bp.before_request
def require_login():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    # نمنع فقط الأساتذة من الدخول، ونسمح لأي إدارة (مهما كان اسم الصلاحية)
    if session.get('role') == 'teacher':
        flash(_("غير مصرح لك بالدخول إلى هذا النظام."), "danger")
        return redirect(url_for('portal'))

# ==========================================
# المسار الرئيسي
# ==========================================
@resit_exams_bp.route('/')
def index():
    db_dict = load_full_db()
    return render_template('resit_exams/index.html', db=db_dict)

# ==========================================
# إدارة البيانات الأساسية
# ==========================================
@resit_exams_bp.route('/manage_data', methods=['GET', 'POST'])
def manage_data():
    db_dict = load_full_db()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # --- الإضافة ---
        if action == 'add_teacher':
            t_input = request.form.get('teachers_input')
            if t_input:
                new_t = list(dict.fromkeys([n.strip() for n in t_input.split('\n') if n.strip()]))
                for t in new_t: add_teacher(t)
                flash(_("تم إضافة {count} أستاذ بنجاح.").format(count=len(new_t)), "success")
                
        elif action == 'add_room':
            r_type = request.form.get('room_type')
            r_input = request.form.get('rooms_input')
            if r_input:
                new_r = list(dict.fromkeys([n.strip() for n in r_input.split('\n') if n.strip()]))
                for r in new_r: add_room(r, r_type)
                flash(_("تم إضافة {count} قاعة بنجاح.").format(count=len(new_r)), "success")
                
        elif action == 'add_level':
            l_input = request.form.get('levels_input')
            if l_input:
                new_l = list(dict.fromkeys([n.strip() for n in l_input.split('\n') if n.strip()]))
                for l in new_l: add_level(l)
                flash(_("تم إضافة {count} مستوى بنجاح.").format(count=len(new_l)), "success")
                
        elif action == 'add_subject':
            level_names = request.form.getlist('level_names')
            subject_names = request.form.get('subject_names', '')
            
            if not level_names:
                flash(_("الرجاء اختيار مستوى واحد على الأقل."), "danger")
            else:
                # دمج المستويات المحددة بفاصل (+)
                combined_level = " + ".join(sorted(level_names))
                
                # 🌟 جلب قاعدة البيانات للتحقق من وجود المستوى المدمج
                db_dict = load_full_db()
                
                # إضافة المستوى المدمج كـ "مستوى جديد" إذا لم يكن موجوداً
                if combined_level not in db_dict.get('levels', []):
                    add_level(combined_level)
                
                # إضافة المواد وربطها بالمستوى المدمج
                for s in subject_names.split('\n'):
                    s = s.strip()
                    if s:
                        add_subject(s, combined_level)
                        
                flash(_("تم إضافة المواد وربطها بالمستويات بنجاح."), "success")

        # --- الحذف (مع التنظيف) ---
        elif action == 'delete_teacher':
            t_del = request.form.get('teacher_name')
            if t_del:
                remove_teacher(t_del)
                ts_dict = db_dict.get('teacher_subjects', {})
                if t_del in ts_dict:
                    del ts_dict[t_del]
                    update_complex_state('teacher_subjects', ts_dict)
                flash(_("تم حذف الأستاذ {t_del} وتنظيف ارتباطاته.").format(t_del=t_del), "danger")

        elif action == 'delete_room':
            r_del = request.form.get('room_name')
            if r_del:
                remove_room(r_del)
                lr_dict = db_dict.get('level_rooms', {})
                for lvl in lr_dict:
                    lr_dict[lvl] = [r for r in lr_dict[lvl] if not r.startswith(r_del)]
                update_complex_state('level_rooms', lr_dict)
                flash(_("تم حذف القاعة {r_del}.").format(r_del=r_del), "danger")

        elif action == 'delete_level':
            l_del = request.form.get('level_name')
            if l_del:
                remove_level(l_del)
                # حذف المواد المرتبطة
                db_dict['subjects'] = [s for s in db_dict['subjects'] if s['level'] != l_del]
                save_full_db(db_dict)
                
                # تنظيف الإسنادات
                ts_dict = db_dict.get('teacher_subjects', {})
                for t in ts_dict:
                    ts_dict[t] = [sub for sub in ts_dict[t] if not sub.endswith(f"({l_del})")]
                update_complex_state('teacher_subjects', ts_dict)
                flash(_("تم حذف المستوى {l_del} وارتباطاته بنجاح.").format(l_del=l_del), "danger")

        elif action == 'delete_subject':
            subject_identifier = request.form.get('subject_identifier')
            if subject_identifier and '|' in subject_identifier:
                s_del, l_for_s = subject_identifier.split('|')
                remove_subject(s_del, l_for_s)
                
                ts_dict = db_dict.get('teacher_subjects', {})
                target_sub = f"{s_del} ({l_for_s})"
                for t in ts_dict:
                    if target_sub in ts_dict[t]:
                        ts_dict[t].remove(target_sub)
                update_complex_state('teacher_subjects', ts_dict)
                flash(_("تم حذف المادة {s_del}.").format(s_del=s_del), "danger")

        # --- التعديل ---
        elif action == 'edit_teacher':
            old_name = request.form.get('old_name')
            new_name = request.form.get('new_name', '').strip()
            if old_name and new_name and old_name != new_name:
                edit_teacher_name(old_name, new_name)
                ts_dict = db_dict.get('teacher_subjects', {})
                if old_name in ts_dict:
                    ts_dict[new_name] = ts_dict.pop(old_name)
                    update_complex_state('teacher_subjects', ts_dict)
                flash(_("تم تعديل الأستاذ إلى [{new_name}].").format(new_name=new_name), "success")

        elif action == 'edit_room':
            old_name = request.form.get('old_name')
            new_name = request.form.get('new_name', '').strip()
            if old_name and new_name and old_name != new_name:
                edit_room_name(old_name, new_name)
                lr_dict = db_dict.get('level_rooms', {})
                modified_lr = False
                for lvl in lr_dict:
                    updated_rooms = []
                    for r in lr_dict[lvl]:
                        if r.startswith(f"{old_name} ("):
                            updated_rooms.append(r.replace(f"{old_name} (", f"{new_name} (", 1))
                            modified_lr = True
                        else:
                            updated_rooms.append(r)
                    lr_dict[lvl] = updated_rooms
                if modified_lr:
                    update_complex_state('level_rooms', lr_dict)
                flash(_("تم تعديل القاعة إلى [{new_name}].").format(new_name=new_name), "success")

        return redirect(url_for('resit_exams.manage_data'))

    return render_template('resit_exams/manage_data.html', db=db_dict)

# ==========================================
# مرحلة إسناد المواد
# ==========================================
@resit_exams_bp.route('/assign_subjects', methods=['GET', 'POST'])
def assign_subjects():
    db_dict = load_full_db()
    ts_dict = db_dict.get('teacher_subjects', {})
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'assign_selected':
            t_name = request.form.get('teacher_name')
            subjects_json = request.form.get('subjects_list')
            
            if t_name and subjects_json:
                subjects_list = json.loads(subjects_json)
                if t_name not in ts_dict:
                    ts_dict[t_name] = []
                
                for s in subjects_list:
                    for other_t in ts_dict:
                        if s in ts_dict[other_t]:
                            ts_dict[other_t].remove(s)
                    if s not in ts_dict[t_name]:
                        ts_dict[t_name].append(s)
                        
                update_complex_state('teacher_subjects', ts_dict)
                flash(_("تم تخصيص المواد للأستاذ {t_name}.").format(t_name=t_name), "success")
                
        elif action == 'unassign_teacher':
            t_name = request.form.get('teacher_name')
            if t_name in ts_dict:
                ts_dict[t_name] = []
                update_complex_state('teacher_subjects', ts_dict)
                flash(_("تم إلغاء مواد الأستاذ {t_name}.").format(t_name=t_name), "warning")
                
        elif action == 'unassign_subject':
            s_name = request.form.get('subject_name')
            for t in ts_dict:
                if s_name in ts_dict[t]:
                    ts_dict[t].remove(s_name)
            update_complex_state('teacher_subjects', ts_dict)
            flash(_("تم إلغاء إسناد المادة {s_name}.").format(s_name=s_name), "warning")
            
        return redirect(url_for('resit_exams.assign_subjects'))

    subject_to_teacher = {}
    for t, subs in ts_dict.items():
        for s in subs:
            subject_to_teacher[s] = t

    return render_template('resit_exams/assign_subjects.html', db=db_dict, subject_to_teacher=subject_to_teacher)

# ==========================================
# تخصيص القاعات
# ==========================================
@resit_exams_bp.route('/assign_rooms', methods=['GET', 'POST'])
def assign_rooms():
    db_dict = load_full_db()
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save_level_rooms':
            new_level_rooms = {}
            for level in db_dict.get('levels', []):
                selected_rooms = request.form.getlist(f'rooms_for_{level}')
                if selected_rooms:
                    new_level_rooms[level] = selected_rooms
                
            update_complex_state('level_rooms', new_level_rooms)
            flash(_("تم حفظ تخصيص القاعات بنجاح!"), "success")
            return redirect(url_for('resit_exams.assign_rooms'))
            
    return render_template('resit_exams/assign_rooms.html', db=db_dict)


import datetime
ARABIC_DAYS = {6: _("الأحد"), 0: _("الإثنين"), 1: _("الثلاثاء"), 2: _("الأربعاء"), 3: _("الخميس"), 4: _("الجمعة"), 5: _("السبت")}

# ==========================================
# أيام وأوقات الامتحان
# ==========================================
@resit_exams_bp.route('/manage_schedule', methods=['GET', 'POST'])
def manage_schedule():
    db_dict = load_full_db()
    schedule = db_dict.get('schedule', {})
    
    # 🌟 التعديل الأول: ترتيب الأيام القديمة (إذا كان لديك أيام غير مرتبة محفوظة مسبقاً)
    if schedule:
        sorted_schedule = {k: schedule[k] for k in sorted(schedule.keys(), key=lambda x: x.split('(')[1].strip(')') if '(' in x else x)}
        if list(schedule.keys()) != list(sorted_schedule.keys()):
            schedule = sorted_schedule
            update_complex_state('schedule', schedule)
            db_dict['schedule'] = schedule
        
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add_day':
            date_str = request.form.get('exam_date')
            if date_str:
                date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                
                # ✨ تعريف القاموس هنا ليترجم اليوم فوراً حسب لغة المستخدم الحالية ✨
                LOCALIZED_DAYS = {
                    6: _("الأحد"), 0: _("الإثنين"), 1: _("الثلاثاء"), 
                    2: _("الأربعاء"), 3: _("الخميس"), 4: _("الجمعة"), 5: _("السبت")
                }
                
                day_name = LOCALIZED_DAYS[date_obj.weekday()]
                new_day = f"{day_name} ({date_str})"
                if new_day not in schedule:
                    schedule[new_day] = []
                    
                    schedule = {k: schedule[k] for k in sorted(schedule.keys(), key=lambda x: x.split('(')[1].strip(')') if '(' in x else x)}
                    
                    update_complex_state('schedule', schedule)
                    flash(_("تم إضافة يوم {new_day}.").format(new_day=new_day), "success")
                    
        elif action == 'delete_day':
            day_to_del = request.form.get('day_key')
            if day_to_del in schedule:
                del schedule[day_to_del]
                update_complex_state('schedule', schedule)
                flash(_("تم حذف يوم {day_to_del}.").format(day_to_del=day_to_del), "danger")
                
        elif action == 'add_time':
            day_key = request.form.get('day_key')
            time_val = request.form.get('time_val')
            if day_key and time_val:
                schedule[day_key].append({
                    "time": time_val, 
                    "primary_levels": [], 
                    "reserve_levels": []
                })
                update_complex_state('schedule', schedule)
                flash(_("تم إضافة الفترة الزمنية."), "success")
                
        elif action == 'delete_time':
            day_key = request.form.get('day_key')
            time_idx = request.form.get('time_idx')
            if day_key in schedule and time_idx is not None:
                try:
                    idx = int(time_idx)
                    del schedule[day_key][idx]
                    if not schedule[day_key]: del schedule[day_key]
                    update_complex_state('schedule', schedule)
                except ValueError: pass
                
        elif action == 'save_levels':
            day_key = request.form.get('day_key')
            if day_key in schedule:
                for idx, slot in enumerate(schedule[day_key]):
                    slot['primary_levels'] = request.form.getlist(f"primary_levels_{idx}")
                    slot['reserve_levels'] = request.form.getlist(f"reserve_levels_{idx}")
                update_complex_state('schedule', schedule)
                flash(_("تم تحديث المستويات في الفترات."), "success")

        # 🌟 مسار تكرار/نسخ اليوم 🌟
        elif action == 'duplicate_day':
            source_day = request.form.get('source_day')
            target_date_str = request.form.get('target_date')
            
            if source_day and target_date_str:
                import copy
                date_obj = datetime.datetime.strptime(target_date_str, '%Y-%m-%d')
                
                # ✨ استخدام القاموس المترجم هنا أيضاً ✨
                LOCALIZED_DAYS = {
                    6: _("الأحد"), 0: _("الإثنين"), 1: _("الثلاثاء"), 
                    2: _("الأربعاء"), 3: _("الخميس"), 4: _("الجمعة"), 5: _("السبت")
                }
                day_name = LOCALIZED_DAYS[date_obj.weekday()]
                target_day_key = f"{day_name} ({target_date_str})"
                
                if source_day in schedule:
                    schedule[target_day_key] = copy.deepcopy(schedule[source_day])
                    
                    schedule = {k: schedule[k] for k in sorted(schedule.keys(), key=lambda x: x.split('(')[1].strip(')') if '(' in x else x)}
                    
                    update_complex_state('schedule', schedule)
                    flash(_("✅ تم نسخ فترات ومستويات يوم [{source_day}] إلى [{target_day_key}] بنجاح.").format(source_day=source_day, target_day_key=target_day_key), "success")

        return redirect(url_for('resit_exams.manage_schedule'))
        
    return render_template('resit_exams/manage_schedule.html', db=db_dict)

# ==========================================
# إدارة القيود والشروط (المسار المفقود)
# ==========================================
@resit_exams_bp.route('/manage_constraints', methods=['GET', 'POST'])
def manage_constraints():
    db_dict = load_full_db()
    
    # تهيئة ذواكر القيود إذا لم تكن موجودة
    # تهيئة ذواكر القيود إذا لم تكن موجودة
    if 'constraints' not in db_dict:
        db_dict['constraints'] = {}
        
    c_db = db_dict['constraints']
    if 'incompatible_levels' not in c_db: c_db['incompatible_levels'] = []
    if 'carpool_pairs' not in c_db: c_db['carpool_pairs'] = []
    if 'conflict_pairs' not in c_db: c_db['conflict_pairs'] = []
    if 'no_first_slot_teachers' not in c_db: c_db['no_first_slot_teachers'] = []

    # 🌟 الهجرة الآلية: تحويل القائمة القديمة إلى قاموس متقدم لدعم الأنماط
    if 'prioritized_teachers' not in c_db: 
        c_db['prioritized_teachers'] = {}
    elif isinstance(c_db['prioritized_teachers'], list):
        c_db['prioritized_teachers'] = {t: 'flexible' for t in c_db['prioritized_teachers']}

    if request.method == 'POST':
        action = request.form.get('action')
        
        # 1. قيد تعارض المستويات
        if action == 'add_incompatible':
            l1, l2 = request.form.get('level1'), request.form.get('level2')
            if l1 and l2 and l1 != l2:
                pair = sorted([l1, l2])
                if pair not in c_db['incompatible_levels']:
                    c_db['incompatible_levels'].append(pair)
                    update_complex_state('constraints', c_db)
                    flash(_("تم إضافة قيد تعارض المستويات."), "success")
        elif action == 'del_incompatible':
            idx = int(request.form.get('idx'))
            c_db['incompatible_levels'].pop(idx)
            update_complex_state('constraints', c_db)
            flash(_("تم حذف القيد."), "danger")
            
        elif action == 'auto_extract_incompatible':
            levels = db_dict.get('levels', [])
            added_count = 0
            for complex_level in levels:
                if '+' in complex_level:
                    sub_levels = [part.strip() for part in complex_level.split('+')]
                    for sub in sub_levels:
                        if sub in levels:
                            pair = sorted([complex_level, sub])
                            if pair not in c_db['incompatible_levels']:
                                c_db['incompatible_levels'].append(pair)
                                added_count += 1
            if added_count > 0:
                update_complex_state('constraints', c_db)
                flash(_("🤖 تم مسح المستويات وتوليد ({count}) قيد تعارض آلياً بنجاح!").format(count=added_count), "success")
            else:
                flash(_("لم يتم العثور على تعارضات جديدة لإضافتها."), "info")
        
        # 2. 🌟 الأساتذة ذوو الأولوية (بالتحديث الجديد - تحديد متعدد)
        elif action == 'add_prioritized':
            selected_teachers = request.form.getlist('teachers')
            pref = request.form.get('pref', 'flexible')
            added_count = 0
            
            if selected_teachers:
                for teacher in selected_teachers:
                    if teacher not in c_db['prioritized_teachers']:
                        c_db['prioritized_teachers'][teacher] = pref
                        added_count += 1
                        
                if added_count > 0:
                    update_complex_state('constraints', c_db)
                    flash(_("تم إضافة ({count}) أساتذة لقائمة الأولوية بنمط ({pref}).").format(count=added_count, pref=pref), "success")
            else:
                flash(_("الرجاء تأشير أستاذ واحد على الأقل من القائمة."), "danger")

        elif action == 'add_all_prioritized':
            pref = request.form.get('pref', 'flexible')
            all_teachers = db_dict.get('teachers', [])
            added_count = 0
            for t in all_teachers:
                if t not in c_db['prioritized_teachers']:
                    c_db['prioritized_teachers'][t] = pref
                    added_count += 1
            
            if added_count > 0:
                update_complex_state('constraints', c_db)
                flash(_("تم إضافة جميع الأساتذة المتبقين وعددهم ({count}) إلى قائمة الأولوية.").format(count=added_count), "success")
            else:
                flash(_("جميع الأساتذة متواجدون بالفعل في قائمة الأولوية."), "info")
            
        elif action == 'del_prioritized':
            teacher = request.form.get('teacher')
            if teacher and teacher in c_db.get('prioritized_teachers', {}):
                del c_db['prioritized_teachers'][teacher]
            update_complex_state('constraints', c_db)
            flash(_("تم حذف الأستاذ من قائمة الأولوية بنجاح."), "danger")
        
        # 3. أساتذة في سيارة واحدة (مرافقة)
        elif action == 'add_carpool':
            t1, t2 = request.form.get('t1'), request.form.get('t2')
            if t1 and t2 and t1 != t2:
                pair = sorted([t1, t2])
                if pair not in c_db['carpool_pairs']:
                    c_db['carpool_pairs'].append(pair)
                    update_complex_state('constraints', c_db)
                    flash(_("تم إضافة قيد المرافقة."), "success")
        elif action == 'del_carpool':
            idx = int(request.form.get('idx'))
            c_db['carpool_pairs'].pop(idx)
            update_complex_state('constraints', c_db)
            flash(_("تم حذف قيد المرافقة."), "danger")
            
        # 4. قيد الانفصال (عدم الاشتراك)
        elif action == 'add_conflict':
            t1, t2 = request.form.get('t1'), request.form.get('t2')
            if t1 and t2 and t1 != t2:
                pair = sorted([t1, t2])
                if pair not in c_db['conflict_pairs']:
                    c_db['conflict_pairs'].append(pair)
                    update_complex_state('constraints', c_db)
                    flash(_("تم إضافة قيد الانفصال."), "success")
        elif action == 'del_conflict':
            idx = int(request.form.get('idx'))
            c_db['conflict_pairs'].pop(idx)
            update_complex_state('constraints', c_db)
            flash(_("تم حذف قيد الانفصال."), "danger")
            
        # 5. إعفاء من الحصة الأولى
        elif action == 'add_no_first':
            t = request.form.get('teacher')
            if t and t not in c_db['no_first_slot_teachers']:
                c_db['no_first_slot_teachers'].append(t)
                update_complex_state('constraints', c_db)
                flash(_("تم إعفاء الأستاذ من الحصة الأولى."), "success")
        elif action == 'del_no_first':
            t = request.form.get('teacher')
            c_db['no_first_slot_teachers'].remove(t)
            update_complex_state('constraints', c_db)
            flash(_("تم إلغاء الإعفاء."), "danger")
            
        # 6. مجموعات العزل
        elif action == 'add_isolation_group':
            group_name = request.form.get('group_name', '').strip()
            group_levels = request.form.getlist('group_levels')
            if group_name and group_levels:
                if 'isolation_groups' not in c_db:
                    c_db['isolation_groups'] = {}
                c_db['isolation_groups'][group_name] = group_levels
                update_complex_state('constraints', c_db)
                flash(_("تم إنشاء مجموعة العزل [{group_name}] وتشفير مستوياتها بنجاح.").format(group_name=group_name), "success")
                
        elif action == 'del_isolation_group':
            group_name = request.form.get('group_name')
            if 'isolation_groups' in c_db and group_name in c_db['isolation_groups']:
                del c_db['isolation_groups'][group_name]
                update_complex_state('constraints', c_db)
                flash(_("تم فك العزل وحذف المجموعة [{group_name}].").format(group_name=group_name), "success")
        
        return redirect(url_for('resit_exams.manage_constraints'))
        
    return render_template('resit_exams/manage_constraints.html', db=db_dict, c_db=c_db)

# ==========================================
# 🌟 الاستيراد المباشر من الامتحانات السداسية
# ==========================================
@resit_exams_bp.route('/import_from_exams', methods=['POST'])
def import_from_exams():
    try:
        # 1. استيراد النماذج (Models) الخاصة بالامتحانات السداسية من قاعدة البيانات
        from app.database import ExamTeacher, ExamRoom, ExamLevel, ExamSubject
        
        tenant_id = session.get('tenant_id')
        if not tenant_id:
            flash(_("غير مصرح لك بإجراء هذه العملية."), "danger")
            return redirect(url_for('resit_exams.manage_data'))
        
        # 2. جلب البيانات الفعلية من جداول الامتحانات السداسية للقسم الحالي
        exams_teachers = ExamTeacher.query.filter_by(tenant_id=tenant_id).all()
        exams_rooms = ExamRoom.query.filter_by(tenant_id=tenant_id).all()
        exams_levels = ExamLevel.query.filter_by(tenant_id=tenant_id).all()
        exams_subjects = ExamSubject.query.filter_by(tenant_id=tenant_id).all()
        
        # التحقق مما إذا كانت هناك بيانات فعلياً
        if not exams_teachers and not exams_subjects:
            flash(_("⚠️ البيانات في نظام الامتحانات السداسية تبدو فارغة. تأكد من إدخالها هناك أولاً."), "warning")
            return redirect(url_for('resit_exams.manage_data'))
            
        # 3. جلب صندوق الاستدراكي الحالي
        resit_db = load_full_db()
        
        # 4. تحويل البيانات العلائقية إلى صيغة القاموس (JSON) الخاصة بالاستدراكي
        
        # أ. الأساتذة
        resit_db['teachers'] = [t.name for t in exams_teachers]
        
        # ب. القاعات
        resit_db['rooms'] = {r.name: r.type for r in exams_rooms}
        
        # ج. المستويات الأساسية
        levels_set = set([l.name for l in exams_levels])
        
        # د. المواد (هنا التعديل الذكي لدمج المستويات)
        resit_db['subjects'] = []
        for s in exams_subjects:
            # قراءة المستويات المتعددة بالهيكل الجديد ودمجها
            levels_list = sorted([l.name for l in s.levels]) if hasattr(s, 'levels') and s.levels else []
            combined_level = " + ".join(levels_list) if levels_list else _("بدون مستوى")
            
            # تسجيل المستوى المدمج كـ "مستوى جديد" في النظام
            if combined_level != _("بدون مستوى"):
                levels_set.add(combined_level)
                
            resit_db['subjects'].append({
                "name": s.name, 
                "level": combined_level
            })
            
        resit_db['levels'] = sorted(list(levels_set))
        
        # هـ. الإسناد (استخراج مواد كل أستاذ بنفس الهيكل المدمج)
        teacher_subjects_dict = {}
        for t in exams_teachers:
            assigned_subs = []
            for s in t.subjects:
                levels_list = sorted([l.name for l in s.levels]) if hasattr(s, 'levels') and s.levels else []
                combined_level = " + ".join(levels_list) if levels_list else _("بدون مستوى")
                assigned_subs.append(f"{s.name} ({combined_level})")
            
            if assigned_subs:
                teacher_subjects_dict[t.name] = assigned_subs
                
        resit_db['teacher_subjects'] = teacher_subjects_dict
        
        # 5. حفظ التغييرات في قاعدة بيانات الاستدراكي
        save_full_db(resit_db)
        
        flash(_("✅ تم سحب البيانات بنجاح! استُورِد: ({t_count}) أستاذ، ({r_count}) قاعة، و({s_count}) مادة.").format(t_count=len(resit_db['teachers']), r_count=len(resit_db['rooms']), s_count=len(resit_db['subjects'])), "success")
        
    except Exception as e:
        import traceback
        traceback.print_exc() # لطباعة الخطأ بدقة في الطرفية إذا حدث
        flash(_("❌ خطأ في الاستيراد: {error}").format(error=str(e)), "danger")
        
    return redirect(url_for('resit_exams.manage_data'))


# ==========================================
# 🌟 مسارات الاستيراد والتصدير المتوافقة مع السحابة (SaaS)
# ==========================================
@resit_exams_bp.route('/export_data')
def export_data():
    """تصدير بيانات الاستدراكي للقسم الحالي على شكل JSON (آمن)"""
    db_dict = load_full_db()
    
    # تحويل القاموس إلى ملف JSON في الذاكرة
    json_data = json.dumps(db_dict, ensure_ascii=False, indent=4)
    buffer = BytesIO()
    buffer.write(json_data.encode('utf-8'))
    buffer.seek(0)
    
    return send_file(
        buffer, 
        as_attachment=True, 
        download_name='resit_exam_backup.json',
        mimetype='application/json'
    )

@resit_exams_bp.route('/import_smart', methods=['POST'])
def import_smart():
    """استيراد وتفريغ البيانات من ملف JSON بأمان تام"""
    if 'data_file' not in request.files:
        return redirect(url_for('resit_exams.manage_data'))
        
    file = request.files['data_file']
    if file.filename.endswith('.json'):
        try:
            # قراءة الملف من الذاكرة دون حفظه في السيرفر
            file_content = file.read().decode('utf-8')
            imported_data = json.loads(file_content)
            
            # حفظ البيانات الكاملة في الصندوق الخاص بالقسم
            save_full_db(imported_data)
            flash(_("✅ تم استعادة بيانات الاستدراكي بنجاح!"), "success")
            
        except Exception as e:
            flash(_("❌ حدث خطأ أثناء الاستيراد: {error}").format(error=str(e)), "danger")
    else:
        flash(_("❌ يرجى رفع ملف بصيغة .json"), "danger")
        
    return redirect(url_for('resit_exams.manage_data'))


from app.resit_exporter import generate_levels_word, generate_teachers_word

# ==========================================
# 🌟 مسارات تشغيل الخوارزمية (SaaS)
# ==========================================
@resit_exams_bp.route('/generate_schedule', methods=['GET'])
def generate_schedule():
    db_dict = load_full_db()
    # جلب المخالفات من التوليد السابق إن وجدت
    app_state = {'is_generated': 'final_schedule' in db_dict, 'violations': db_dict.get('final_violations', [])}
    return render_template('resit_exams/generate_schedule.html', db=db_dict, app_state=app_state)

# ==========================================
# 🌟 مسارات تشغيل الخوارزمية (SaaS & Desktop)
# ==========================================
@resit_exams_bp.route('/start_solver', methods=['POST'])
def start_solver():
    from app.resit_tasks import run_resit_distribution_task
    
    data = request.get_json(silent=True) or request.form or {}
    algo_choice = data.get('algo_choice', 'lns')
    strategy = data.get('strategy', 'teacher')
    
    try:
        duration = int(float(data.get('lns_duration') or data.get('duration') or 10))
        destruction_rate = int(float(data.get('lns_destruction') or data.get('destruction_rate') or 20))
    except (ValueError, TypeError):
        duration, destruction_rate = 10, 20
        
    db_dict = load_full_db()
    constraints = db_dict.get('constraints', {})
    constraints.update({'algo_choice': algo_choice, 'strategy': strategy, 'duration': duration, 'destruction_rate': destruction_rate})
    update_complex_state('constraints', constraints)
    
    tenant_id = session.get('tenant_id')
    
    # 🚀 المحول الذكي
    from flask import current_app
    import threading
    mode = current_app.config.get('APP_MODE')
    
    if mode == 'desktop':
        # تصفير الذاكرة الحية
        from app.memory_logger import store
        with store.lock:
            store.status[f"running_{tenant_id}"] = True
            store.status[f"done_{tenant_id}"] = None
            store.status[f"progress_{tenant_id}"] = None
            
        app_obj = current_app._get_current_object()
        def run_thread():
            with app_obj.app_context():
                from app.resit_tasks import execute_resit_distribution
                execute_resit_distribution(tenant_id, algo_choice, duration, destruction_rate, strategy, celery_task=None)
        threading.Thread(target=run_thread).start()
    else:
        # مسار السحابة المعتاد
        task = run_resit_distribution_task.delay(tenant_id, algo_choice, duration, destruction_rate, strategy)
        session['resit_task_id'] = task.id
    
    return jsonify({"status": "started", "duration_used": duration})


@resit_exams_bp.route('/solver_progress')
def get_solver_progress():
    tenant_id = session.get('tenant_id')
    from flask import current_app
    mode = current_app.config.get('APP_MODE')
    
    if mode == 'desktop':
        # جلب حالة التقدم من الذاكرة الحية (RAM)
        from app.memory_logger import store
        with store.lock:
            is_running = store.status.get(f"running_{tenant_id}", False)
            done_violations = store.status.get(f"done_{tenant_id}")
            progress = store.status.get(f"progress_{tenant_id}")
            
        if done_violations is not None:
            return jsonify({"is_running": False, "done": True, "violations": done_violations})
        elif progress:
            return jsonify({
                "is_running": True, "done": False,
                "elapsed": progress.get('elapsed', 0),
                "duration": progress.get('duration', 10),
                "unassigned": progress.get('unassigned', 0),
                "hard": progress.get('hard', 0),
                "soft": progress.get('soft', 0)
            })
        elif is_running:
            return jsonify({"is_running": True, "elapsed": 0, "done": False})
        else:
            return jsonify({"is_running": False, "done": False})
            
    else:
        # جلب حالة التقدم من خادم Celery
        from app.resit_tasks import run_resit_distribution_task
        task_id = session.get('resit_task_id')
        if not task_id:
            return jsonify({"is_running": False, "done": False})
            
        task = run_resit_distribution_task.AsyncResult(task_id)
        if task.state == 'PENDING':
            response = {"is_running": True, "elapsed": 0, "done": False}
        elif task.state == 'PROGRESS':
            response = {
                "is_running": True, "done": False,
                "elapsed": task.info.get('elapsed', 0),
                "duration": task.info.get('duration', 10),
                "unassigned": task.info.get('unassigned', 0),
                "hard": task.info.get('hard', 0),
                "soft": task.info.get('soft', 0)
            }
        elif task.state == 'SUCCESS':
            response = {
                "is_running": False, 
                "done": True,
                "violations": task.info.get('violations', []) if isinstance(task.info, dict) else []
            }
        else:
            response = {"is_running": False, "done": True, "error": str(task.info)}
            
        return jsonify(response)
    
# ==========================================
# 🌟 مسارات تحميل ملفات Word
# ==========================================
@resit_exams_bp.route('/download/<doc_type>')
def download_doc(doc_type):
    db_dict = load_full_db()
    final_schedule = db_dict.get('final_schedule')
        
    if not final_schedule:
        flash(_("لم يتم توليد الجدول بعد."), "danger")
        return redirect(url_for('resit_exams.generate_schedule'))
        
    # ✨ التقاط اللغة المختارة من الرابط (الافتراضي عربي)
    lang = request.args.get('lang', 'ar')
        
    if doc_type == 'levels':
        doc_stream = generate_levels_word(db_dict, final_schedule, lang=lang)
        if lang == 'en':
            filename = "Resit_Levels.docx"
        elif lang == 'fr':
            filename = "Rattrapage_Niveaux.docx"
        else:
            filename = _("جداول_المستويات_استدراكي.docx")
            
    elif doc_type == 'teachers':
        doc_stream = generate_teachers_word(db_dict, final_schedule, lang=lang)
        if lang == 'en':
            filename = "Resit_Teachers.docx"
        elif lang == 'fr':
            filename = "Rattrapage_Profs.docx"
        else:
            filename = _("جداول_الأساتذة_استدراكي.docx")
    else:
        return redirect(url_for('resit_exams.generate_schedule'))
        
    doc_stream.seek(0)
    return send_file(
        doc_stream, 
        as_attachment=True, 
        download_name=filename, 
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )

# ==========================================
# 🌟 المرحلة 7: التصدير للإكسل، الاستيراد، والنشر للأساتذة
# ==========================================
@resit_exams_bp.route('/phase7', methods=['GET'])
def phase7():
    db_dict = load_full_db()
    final_schedule = db_dict.get('final_schedule', {})
    
    # التحقق من حالة النشر الحالية
    tenant_id = session.get('tenant_id')
    from app.database import ExamSetting
    pub_setting = ExamSetting.query.filter_by(key='is_resit_published', tenant_id=tenant_id).first()
    is_published = pub_setting.value == '1' if pub_setting else False
    
    return render_template('resit_exams/phase7.html', db=db_dict, has_schedule=bool(final_schedule), is_published=is_published)

# ==========================================
# 🌟 التصدير للإكسل (ديناميكي يدعم العربية والإنجليزية)
# ==========================================
@resit_exams_bp.route('/export_excel')
def export_excel():
    db_dict = load_full_db()
    final_schedule = db_dict.get('final_schedule', {})
    if not final_schedule:
        flash(_("لا يوجد جدول لتصديره."), "danger")
        return redirect(url_for('resit_exams.phase7'))

    import pandas as pd
    import io
    import re
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    # ✨ 1. تحديد لغة الجلسة الحالية واتجاه الملف
    lang = session.get('lang', 'ar')
    is_rtl = (lang == 'ar')

    # استخراج جميع الأيام، الفترات، والمستويات (مع ترتيب الأيام ذكياً حسب التاريخ)
    all_days = sorted(list(final_schedule.keys()), key=lambda x: x.split('(')[1].strip(')') if '(' in x else x)
    all_times = set()
    all_levels = set()

    for day, times in final_schedule.items():
        for time_val, levels in times.items():
            all_times.add(time_val)
            for level in levels.keys():
                all_levels.add(level)

    all_times = sorted(list(all_times))
    all_levels = sorted(list(all_levels))

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # بناء الجداول بحيث يكون كل مستوى في Sheet مستقل
    for level_name in all_levels:
        df_level = pd.DataFrame(index=all_times, columns=all_days)
        df_level.index.name = _("الفترة")

        for day in all_days:
            for time_val in all_times:
                if day in final_schedule and time_val in final_schedule[day]:
                    levels_dict = final_schedule[day][time_val]
                    if level_name in levels_dict:
                        data = levels_dict[level_name]
                        subject = data.get("subject", "")
                        teachers = "، ".join(data.get("subject_teachers", []))
                        rooms = data.get("rooms", {})

                        halls_list = list(rooms.keys())
                        halls_str = "، ".join(halls_list)

                        # تنسيق الحراس لربطهم بقاعاتهم بشكل مقروء للتعديل
                        guards_details = []
                        for r_name, g_list in rooms.items():
                            g_str = "، ".join(g_list) if g_list else _("بدون حراس")
                            guards_details.append(f"{r_name}: {g_str}")
                        guards_str = " | ".join(guards_details)

                        # محاكاة نفس طريقة الامتحانات العادية تماماً
                        cell_content = f"{subject}\n::: {teachers}\n::: {level_name}\n::: {halls_str}\n::: {guards_str}"

                        existing = df_level.at[time_val, day]
                        if pd.notna(existing) and str(existing).strip() != '':
                            df_level.at[time_val, day] = str(existing) + "\n\n====================\n\n" + cell_content
                        else:
                            df_level.at[time_val, day] = cell_content

        safe_sheet_name = re.sub(r'[\\*?:/\[\]]', '-', level_name)[:31]
        df_level.to_excel(writer, sheet_name=safe_sheet_name)

        # التنسيقات الجمالية
        worksheet = writer.sheets[safe_sheet_name]
        
        # ✨ 2. تفعيل اتجاه الورقة حسب اللغة (يمين لليسار للعربية فقط)
        worksheet.sheet_view.rightToLeft = is_rtl
        
        worksheet.column_dimensions['A'].width = 18
        for i in range(2, len(all_days) + 2):
            worksheet.column_dimensions[get_column_letter(i)].width = 30

        # ✨ 3. تحديد محاذاة النص بناءً على اللغة
        align_horizontal = 'right' if is_rtl else 'left'
        reading_order = 2 if is_rtl else 1 # 2=RTL, 1=LTR
        
        wrap_alignment = Alignment(wrap_text=True, horizontal=align_horizontal, vertical='center', readingOrder=reading_order)
        
        for row in worksheet.iter_rows():
            if row[0].row == 1:
                worksheet.row_dimensions[row[0].row].height = 35
            else:
                worksheet.row_dimensions[row[0].row].height = None

            for cell in row:
                cell.alignment = wrap_alignment

    writer.close()
    output.seek(0)
    
    # ✨ 4. تسمية الملف بناءً على اللغة الحالية
    file_name = "Resit_Schedule_Edit.xlsx" if lang == 'en' else _("الجدول_الاستدراكي_للتعديل.xlsx")
    
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==========================================
# 🌟 استيراد الجدول من الإكسل (فهم نفس الهيكل الذكي)
# ==========================================
@resit_exams_bp.route('/import_excel', methods=['POST'])
def import_excel():
    if 'excel_file' not in request.files:
        return redirect(url_for('resit_exams.phase7'))

    file = request.files['excel_file']
    if file.filename.endswith('.xlsx'):
        try:
            import pandas as pd
            db_dict = load_full_db()
            new_schedule = {}

            xls = pd.read_excel(file, sheet_name=None, index_col=0, dtype=str)

            for sheet_name, df in xls.items():
                for day in df.columns:
                    for time_val in df.index:
                        cell_value = df.at[time_val, day]
                        if pd.notna(cell_value):
                            cell_str = str(cell_value).strip()

                            # التوافق مع الفواصل الذكية
                            if "====================" in cell_str:
                                exams_in_cell = cell_str.split('\n\n====================\n\n')
                            elif "\n:::" in cell_str:
                                exams_in_cell = [cell_str]
                            else:
                                exams_in_cell = cell_str.split('\n')

                            for exam_block in exams_in_cell:
                                clean_block = exam_block.replace('\n', ' ')
                                if ':::' in clean_block:
                                    parts = [part.strip() for part in clean_block.split(':::')]
                                    if len(parts) >= 5:
                                        subject = parts[0]
                                        teachers = [t.strip() for t in parts[1].split('،') if t.strip()]
                                        level = parts[2]
                                        guards_part = parts[4]

                                        if day not in new_schedule: new_schedule[day] = {}
                                        if time_val not in new_schedule[day]: new_schedule[day][time_val] = {}
                                        if level not in new_schedule[day][time_val]:
                                            new_schedule[day][time_val][level] = {"subject": subject, "subject_teachers": teachers, "rooms": {}}

                                        # إعادة قراءة القاعات وحراسها بدقة
                                        room_blocks = guards_part.split('|')
                                        for r_block in room_blocks:
                                            if ':' in r_block:
                                                r_name, r_guards = r_block.split(':', 1)
                                                r_name = r_name.strip()
                                                guards_list = [g.strip() for g in r_guards.split('،') if g.strip() and g.strip() != _('بدون حراس')]
                                                new_schedule[day][time_val][level]["rooms"][r_name] = guards_list

            update_complex_state('final_schedule', new_schedule)
            flash(_("✅ تم استيراد التعديلات من الإكسل بنجاح وتحديث النظام!"), "success")

        except Exception as e:
            import traceback; traceback.print_exc()
            flash(_("❌ خطأ في قراءة ملف الإكسل: التفاصيل: {error}").format(error=str(e)), "danger")
    else:
        flash(_("يرجى رفع ملف بصيغة .xlsx"), "danger")

    return redirect(url_for('resit_exams.phase7'))

@resit_exams_bp.route('/publish_resit', methods=['POST'])
def publish_resit():
    tenant_id = session.get('tenant_id')
    db_dict = load_full_db()
    final_schedule = db_dict.get('final_schedule')
    
    if not final_schedule:
        flash(_("لا يوجد جدول لنشره."), "danger")
        return redirect(url_for('resit_exams.phase7'))

    from app.database import ExamSetting, db
    action = request.form.get('action')

    # تجهيز متغيرات قاعدة البيانات
    pub_setting = ExamSetting.query.filter_by(key='is_resit_published', tenant_id=tenant_id).first()
    if not pub_setting:
        pub_setting = ExamSetting(key='is_resit_published', value='0', tenant_id=tenant_id)
        db.session.add(pub_setting)

    sched_setting = ExamSetting.query.filter_by(key='published_resit_schedule', tenant_id=tenant_id).first()
    if not sched_setting:
        sched_setting = ExamSetting(key='published_resit_schedule', value='{}', tenant_id=tenant_id)
        db.session.add(sched_setting)

    if action == 'publish':
        import json
        pub_setting.value = '1'
        sched_setting.value = json.dumps(final_schedule)
        flash(_("✅ تم نشر الجدول الاستدراكي بنجاح! سيظهر الآن في بوابة الأساتذة."), "success")
    else:
        pub_setting.value = '0'
        flash(_("⚠️ تم سحب وإخفاء الجدول الاستدراكي من بوابة الأساتذة."), "warning")

    db.session.commit()
    return redirect(url_for('resit_exams.phase7'))

