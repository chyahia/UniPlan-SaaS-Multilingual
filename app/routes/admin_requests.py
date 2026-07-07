from flask import Blueprint, jsonify, session, request, send_file
from app.database import db, TeacherRequest, Teacher, Course, Setting
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

admin_requests_bp = Blueprint('admin_requests', __name__)

# 1. جلب كل الرغبات وتقرير التضاربات لعرضها للمدير
@admin_requests_bp.route('/api/admin/requests', methods=['GET'])
def get_requests():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    
    # جلب الرغبات، المواد، والأساتذة للقسم الحالي فقط (SaaS)
    reqs = TeacherRequest.query.filter_by(tenant_id=tenant_id).all()
    courses_dict = {c.id: c.name for c in Course.query.filter_by(tenant_id=tenant_id).all()}
    teachers_dict = {t.id: t.name for t in Teacher.query.filter_by(tenant_id=tenant_id).all()}
    
    result = []
    for r in reqs:
        course_ids = json.loads(r.requested_courses) if r.requested_courses else []
        course_names = [courses_dict.get(int(cid), "مادة محذوفة") for cid in course_ids]
        result.append({
            "teacher_id": r.teacher_id,
            "teacher_name": teachers_dict.get(r.teacher_id, "مجهول"),
            "courses": course_names,
            "days": json.loads(r.requested_days) if r.requested_days else [],
            "status": r.status
        })

    # جلب تقرير التضاربات المحفوظ من قاعدة البيانات
    report_setting = Setting.query.filter_by(key='conflict_report', tenant_id=tenant_id).first()
    report = json.loads(report_setting.value) if report_setting and report_setting.value else []

    return jsonify({"requests": result, "report": report})

# 2. اعتماد طلب الأستاذ وتطبيق البيانات
@admin_requests_bp.route('/api/admin/requests/approve/<int:teacher_id>', methods=['POST'])
def approve_request(teacher_id):
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    req = TeacherRequest.query.filter_by(teacher_id=teacher_id, tenant_id=tenant_id).first()
    
    if not req:
        return jsonify({"error": "الطلب غير موجود"}), 404

    course_ids = json.loads(req.requested_courses) if req.requested_courses else []
    days = json.loads(req.requested_days) if req.requested_days else []

    # أ) تحديث جدول المواد (إسناد المواد للأستاذ)
    if course_ids:
        Course.query.filter(Course.id.in_(course_ids), Course.tenant_id == tenant_id).update({"teacher_id": teacher_id}, synchronize_session=False)

    # ب) تحديث جدول القيود (تحديد أيام العمل)
    setting = Setting.query.filter_by(key='schedule_conditions', tenant_id=tenant_id).first()
    conditions = json.loads(setting.value) if setting and setting.value else {}

    if 'teacher_rules' not in conditions: conditions['teacher_rules'] = {}
    
    t_id_str = str(teacher_id)
    if t_id_str not in conditions['teacher_rules']:
        conditions['teacher_rules'][t_id_str] = {"days": days, "limits": [], "rule": "unspecified"}
    else:
        conditions['teacher_rules'][t_id_str]['days'] = days

    if setting:
        setting.value = json.dumps(conditions)
    else:
        new_setting = Setting(key='schedule_conditions', value=json.dumps(conditions), tenant_id=tenant_id)
        db.session.add(new_setting)

    # ج) تغيير حالة الطلب إلى "معتمد"
    req.status = 'معتمد'
    db.session.commit()
    
    return jsonify({"success": True, "message": "✅ تم اعتماد الرغبات وتطبيقها على المرحلتين 3 و 5 بنجاح!"})

# 3. حذف الطلب (الرفض)
@admin_requests_bp.route('/api/admin/requests/reject/<int:teacher_id>', methods=['POST'])
def reject_request(teacher_id):
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    tenant_id = session.get('tenant_id')
    TeacherRequest.query.filter_by(teacher_id=teacher_id, tenant_id=tenant_id).delete()
    db.session.commit()
    return jsonify({"success": True})

# 4. الاعتماد الشامل الآلي (فقط للطلبات الجديدة/المعدلة)
@admin_requests_bp.route('/api/admin/requests/bulk_approve', methods=['POST'])
def bulk_approve_requests():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    
    # جلب الطلبات (الجديدة أو المعدلة فقط)
    reqs = TeacherRequest.query.filter_by(status='قيد المراجعة', tenant_id=tenant_id).all()
    
    courses = Course.query.filter_by(tenant_id=tenant_id).all()
    course_status = {c.id: {'name': c.name, 'teacher_id': c.teacher_id, 'obj': c} for c in courses}
    teachers_dict = {t.id: t.name for t in Teacher.query.filter_by(tenant_id=tenant_id).all()}
    
    # جلب التقرير القديم
    report_setting = Setting.query.filter_by(key='conflict_report', tenant_id=tenant_id).first()
    report = json.loads(report_setting.value) if report_setting and report_setting.value else []
    
    # جلب القيود
    setting = Setting.query.filter_by(key='schedule_conditions', tenant_id=tenant_id).first()
    conditions = json.loads(setting.value) if setting and setting.value else {}
    if 'teacher_rules' not in conditions: conditions['teacher_rules'] = {}

    for req in reqs:
        t_id = req.teacher_id
        t_name = teachers_dict.get(t_id, "مجهول")
        requested_cids = json.loads(req.requested_courses) if req.requested_courses else []
        requested_days = json.loads(req.requested_days) if req.requested_days else []
        
        t_id_str = str(t_id)
        if t_id_str not in conditions['teacher_rules']:
            conditions['teacher_rules'][t_id_str] = {"days": requested_days, "limits": [], "rule": "unspecified"}
        else:
            conditions['teacher_rules'][t_id_str]['days'] = requested_days

        for cid_str in requested_cids:
            cid = int(cid_str)
            if cid in course_status:
                c_info = course_status[cid]
                if c_info['teacher_id'] is None:
                    # إسناد المادة
                    c_info['obj'].teacher_id = t_id
                    c_info['teacher_id'] = t_id
                else:
                    # تسجيل التضارب
                    assigned_to_id = c_info['teacher_id']
                    assigned_to_name = teachers_dict.get(assigned_to_id, "أستاذ آخر")
                    if assigned_to_id != t_id:
                        course_exists = False
                        for r in report:
                            if r['course_name'] == c_info['name']:
                                if t_name not in r['requested_by']:
                                    r['requested_by'] += f" ، {t_name}"
                                course_exists = True
                                break
                        
                        if not course_exists:
                            report.append({
                                "course_name": c_info['name'],
                                "requested_by": t_name,
                                "assigned_to": assigned_to_name
                            })
                    
        # تغيير حالة الطلب
        req.status = 'معتمد'
        
    # حفظ الإعدادات
    if report_setting: 
        report_setting.value = json.dumps(report)
    else: 
        db.session.add(Setting(key='conflict_report', value=json.dumps(report), tenant_id=tenant_id))

    if setting: 
        setting.value = json.dumps(conditions)
    else: 
        db.session.add(Setting(key='schedule_conditions', value=json.dumps(conditions), tenant_id=tenant_id))
        
    db.session.commit()
    
    return jsonify({"success": True, "report": report})

# 5. نشر الجداول النهائية للأساتذة في بواباتهم
@admin_requests_bp.route('/api/admin/publish_schedule', methods=['POST'])
def publish_schedule():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    tenant_id = session.get('tenant_id')
    data = request.json
    prof_schedules = data.get('prof_schedules', {})
    
    # 1. حفظ الجداول
    pub_sched_setting = Setting.query.filter_by(key='published_schedule', tenant_id=tenant_id).first()
    if pub_sched_setting: pub_sched_setting.value = json.dumps(prof_schedules)
    else: db.session.add(Setting(key='published_schedule', value=json.dumps(prof_schedules), tenant_id=tenant_id))
    
    # 2. تفعيل مفتاح النشر
    is_pub_setting = Setting.query.filter_by(key='is_published', tenant_id=tenant_id).first()
    if is_pub_setting: is_pub_setting.value = 'true'
    else: db.session.add(Setting(key='is_published', value='true', tenant_id=tenant_id))
    
    db.session.commit()
    return jsonify({"success": True, "message": "✅ تم نشر الجداول! ستظهر الآن فوراً في حسابات الأساتذة."})

# 6. إلغاء النشر
@admin_requests_bp.route('/api/admin/unpublish_schedule', methods=['POST'])
def unpublish_schedule():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    tenant_id = session.get('tenant_id')
    is_pub_setting = Setting.query.filter_by(key='is_published', tenant_id=tenant_id).first()
    
    if is_pub_setting: 
        is_pub_setting.value = 'false'
    else: 
        db.session.add(Setting(key='is_published', value='false', tenant_id=tenant_id))
        
    db.session.commit()
    return jsonify({"success": True, "message": "🚫 تم سحب الجداول وإخفاؤها عن الأساتذة بنجاح."})

# مسار مسح جميع رغبات الأساتذة دفعة واحدة
@admin_requests_bp.route('/api/admin/requests/delete_all', methods=['POST'])
def delete_all_requests():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    try:
        tenant_id = session.get('tenant_id')
        TeacherRequest.query.filter_by(tenant_id=tenant_id).delete()
        db.session.commit()
        return jsonify({"success": True, "message": "تم مسح جميع الرغبات بنجاح."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    
# مسار مخصص ومضمون لمسح تقرير التضاربات بالكامل
@admin_requests_bp.route('/api/admin/report/clear', methods=['POST'])
def clear_conflict_report():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    try:
        tenant_id = session.get('tenant_id')
        report_setting = Setting.query.filter_by(key='conflict_report', tenant_id=tenant_id).first()
        if report_setting: report_setting.value = '[]'
        else: db.session.add(Setting(key='conflict_report', value='[]', tenant_id=tenant_id))
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# ==================== تصدير التبويبات إلى إكسل المنسق ====================

# 1. تصدير قائمة الرغبات
@admin_requests_bp.route('/api/admin/export_requests_excel', methods=['GET'])
def export_requests_excel():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    reqs = TeacherRequest.query.filter_by(tenant_id=tenant_id).all()
    courses_dict = {c.id: c.name for c in Course.query.filter_by(tenant_id=tenant_id).all()}
    teachers_dict = {t.id: t.name for t in Teacher.query.filter_by(tenant_id=tenant_id).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قائمة الرغبات"
    ws.sheet_view.rightToLeft = True

    headers = ["اسم الأستاذ", "حالة الطلب", "أيام العمل المطلوبة", "المواد المطلوبة"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2c3e50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    for r in reqs:
        course_ids = json.loads(r.requested_courses) if r.requested_courses else []
        course_names = [courses_dict.get(int(cid), "مادة محذوفة") for cid in course_ids]
        days = json.loads(r.requested_days) if r.requested_days else []
        
        status_text = "جديد / معدّل" if r.status == 'قيد المراجعة' else "تمت معالجته"
        courses_stacked = "\n".join([f"• {name}" for name in course_names])
        days_text = " ، ".join(days)
        
        ws.append([
            teachers_dict.get(r.teacher_id, "مجهول"),
            status_text,
            days_text,
            courses_stacked
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center", readingOrder=2)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name="Teacher_Requests.xlsx", as_attachment=True)


# 2. تصدير تقرير التضاربات
@admin_requests_bp.route('/api/admin/export_report_excel', methods=['GET'])
def export_report_excel():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    report_setting = Setting.query.filter_by(key='conflict_report', tenant_id=tenant_id).first()
    report = json.loads(report_setting.value) if report_setting and report_setting.value else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير التضاربات"
    ws.sheet_view.rightToLeft = True

    headers = ["المادة المطلوبة", "طُلبَت من طرف", "تم إسنادها مسبقاً إلى"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="c0392b", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    for r in report:
        ws.append([
            r.get('course_name', ''),
            r.get('requested_by', ''),
            r.get('assigned_to', '')
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center", readingOrder=2)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name="Conflict_Report.xlsx", as_attachment=True)

# ==================== إدارة صلاحيات رؤية المواد للأساتذة ====================

@admin_requests_bp.route('/api/admin/teachers/visibility', methods=['GET'])
def get_teachers_visibility():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    teachers = Teacher.query.filter_by(tenant_id=tenant_id).order_by(Teacher.name).all()
    
    return jsonify([{"id": t.id, "name": t.name, "show_assigned": getattr(t, 'show_assigned', 0)} for t in teachers])

@admin_requests_bp.route('/api/admin/teachers/visibility', methods=['POST'])
def update_teachers_visibility():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    data = request.json
    teacher_ids = data.get('teacher_ids', [])
    status = data.get('status', 0)
    all_teachers = data.get('all_teachers', False)
    tenant_id = session.get('tenant_id')
    
    if all_teachers:
        Teacher.query.filter_by(tenant_id=tenant_id).update({"show_assigned": status})
    else:
        Teacher.query.filter(Teacher.id.in_(teacher_ids), Teacher.tenant_id == tenant_id).update({"show_assigned": status}, synchronize_session=False)
        
    db.session.commit()
    return jsonify({"success": True, "message": "✅ تم تحديث صلاحيات الرؤية بنجاح!"})

# ==================== التحكم في قفل وفتح الرغبات ====================
@admin_requests_bp.route('/api/admin/requests_lock', methods=['GET', 'POST'])
def handle_requests_lock():
    if session.get('role') not in ['super_admin', 'tenant_admin']: return jsonify({"error": "غير مصرح"}), 403
    
    tenant_id = session.get('tenant_id')
    
    if request.method == 'POST':
        data = request.json
        is_locked = 'true' if data.get('locked') else 'false'
        
        setting = Setting.query.filter_by(key='requests_locked', tenant_id=tenant_id).first()
        if setting: setting.value = is_locked
        else: db.session.add(Setting(key='requests_locked', value=is_locked, tenant_id=tenant_id))
            
        db.session.commit()
        return jsonify({"success": True, "locked": is_locked == 'true'})
    else:
        setting = Setting.query.filter_by(key='requests_locked', tenant_id=tenant_id).first()
        is_locked = setting.value == 'true' if setting else False
        return jsonify({"locked": is_locked})