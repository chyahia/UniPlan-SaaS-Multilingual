from flask import Blueprint, request, jsonify, Response, session
import time
import json
import traceback
import copy
import io
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font

# استدعاء الخوارزميات (بدون المتغيرات العامة القديمة)
from app.services.algorithms import (
    run_tabu_search, run_large_neighborhood_search, run_variable_neighborhood_search, 
    run_greedy_search_for_best_result, refine_and_compact_schedule
)

from app.database import db, Teacher, Room, Level, Course, Setting

# ✨ استدعاء أدوات السحابة الجديدة
from app.celery_setup import celery_app
from app.redis_logger import RedisLogQueue, redis_client

generation_bp = Blueprint('generation', __name__)

# ✨ كائن ذكي (Proxy) يخدع الخوارزمية ويجعلها تقرأ أوامر التوقف والطفرة من Redis مباشرة!
class SchedulingStateProxy:
    def __init__(self, tenant_id):
        self.tenant_id = tenant_id
        self.log_q = RedisLogQueue(tenant_id)
        self.redis = redis_client

    def get(self, key, default=None):
        if key == 'should_stop': 
            return self.log_q.should_stop()
        if key == 'force_mutation': 
            return self.redis.get(f"mutation:tenant_{self.tenant_id}") is not None
        if key == 'mutation_intensity':
            val = self.redis.get(f"mutation:tenant_{self.tenant_id}")
            return int(val) if val else default
        return default

    def __setitem__(self, key, value): 
        # تستخدمه الخوارزمية لإطفاء زر الطفرة بعد استعمالها
        if key == 'force_mutation' and not value:
            self.redis.delete(f"mutation:tenant_{self.tenant_id}")
            
    def pop(self, key, default=None): 
        if key == 'mutation_intensity':
            self.redis.delete(f"mutation:tenant_{self.tenant_id}")

    def __contains__(self, key):
        return key in ['should_stop', 'force_mutation', 'mutation_intensity']


# ================= مسارات الواجهة الأمامية (Flask API) =================

@generation_bp.route('/api/generate', methods=['POST'])
def generate_schedule():
    tenant_id = session.get('tenant_id')
    log_q = RedisLogQueue(tenant_id)
    
    if log_q.is_running():
        return jsonify({"success": False, "error": "عملية التوزيع تعمل حالياً في قسمك."}), 400

    data = request.json
    # تهيئة الشاشة السوداء في Redis
    log_q.clear_logs()
    log_q.set_running(True)
    log_q.set_stop_flag(False)
    redis_client.delete(f"mutation:tenant_{tenant_id}") # تنظيف أي طفرات سابقة

    # ✨ إرسال المهمة للطباخ (Celery) للعمل في الخلفية باستخدام delay
    background_generation_task.delay(
        tenant_id, 
        data.get('strict_hierarchy'), 
        data.get('algorithms'), 
        data.get('settings', {})
    )
    return jsonify({"success": True})


@generation_bp.route('/api/refine', methods=['POST'])
def start_refinement():
    tenant_id = session.get('tenant_id')
    log_q = RedisLogQueue(tenant_id)
    
    if log_q.is_running():
        return jsonify({"error": "هناك عملية قيد التشغيل بالفعل"}), 400
    
    data = request.json
    current_schedule = data.get('schedule')
    if not current_schedule: return jsonify({"error": "الجدول غير موجود أو فارغ."}), 400
        
    log_q.clear_logs()
    log_q.set_running(True)
    log_q.set_stop_flag(False)

    background_refinement_task.delay(
        tenant_id, 
        current_schedule, 
        data.get('level', 'balanced'), 
        data.get('teachers', [])
    )
    return jsonify({"success": True})


@generation_bp.route('/api/stream_logs')
@generation_bp.route('/stream-logs', methods=['GET'])
def stream_logs():
    tenant_id = session.get('tenant_id')
    log_q = RedisLogQueue(tenant_id)
    
    def generate():
        last_idx = 0
        while True:
            logs = log_q.get_logs(start_index=last_idx)
            for msg in logs:
                yield f"data: {msg}\n\n"
            last_idx += len(logs)
            
            if not log_q.is_running() and last_idx > 0:
                break
            time.sleep(0.5)
            
    return Response(generate(), mimetype='text/event-stream', headers={'X-Accel-Buffering': 'no', 'Cache-Control': 'no-cache'})


@generation_bp.route('/api/stop-generation', methods=['POST'])
def stop_generation():
    tenant_id = session.get('tenant_id')
    log_q = RedisLogQueue(tenant_id)
    log_q.set_stop_flag(True)
    return jsonify({"success": True})


@generation_bp.route('/api/generate/force_mutation', methods=['POST'])
def force_mutation_route():
    tenant_id = session.get('tenant_id')
    data = request.get_json() or {}
    intensity = data.get('intensity', 4) 
    # إرسال أمر الطفرة عبر Redis
    redis_client.set(f"mutation:tenant_{tenant_id}", intensity, ex=3600)
    return jsonify({"success": True})


@generation_bp.route('/api/force_reset', methods=['POST'])
def force_reset():
    tenant_id = session.get('tenant_id')
    log_q = RedisLogQueue(tenant_id)
    log_q.set_running(False)
    log_q.set_stop_flag(True)
    log_q.put("🛑 تم فرض إيقاف الخادم وإعادة التهيئة يدوياً بواسطة المستخدم.")
    return jsonify({"success": True})


# ================= مسارات التصدير والاستيراد (Excel) =================

@generation_bp.route('/api/export_excel', methods=['POST'])
def export_excel():
    data = request.json
    schedule = data.get('schedule', {})
    days = data.get('days', [])
    slots = data.get('slots', [])
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) 
    
    for level, grid in schedule.items():
        safe_title = level.replace("/", "-").replace("\\", "-")[:31]
        ws = wb.create_sheet(title=safe_title)
        ws.sheet_view.rightToLeft = True
        
        ws['A1'] = level
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="34495e", fill_type="solid")
        
        headers = ["الوقت"] + days
        ws.append(headers)
        for cell in ws[2]: 
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ecf0f1", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for slot_idx, slot_name in enumerate(slots):
            row = [slot_name]
            for day_idx, day_name in enumerate(days):
                cell_data = grid[day_idx][slot_idx] if day_idx < len(grid) and slot_idx < len(grid[day_idx]) else []
                if not cell_data:
                    row.append("-")
                else:
                    parts = []
                    for lec in cell_data:
                        parts.append(f"مادة: {lec.get('name','')}\nأستاذ: {lec.get('teacher_name','')}\nقاعة: {lec.get('room','')}")
                    row.append("\n===\n".join(parts))
            ws.append(row)
            
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 30
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return Response(
        out, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=schedule_export.xlsx"}
    )

@generation_bp.route('/api/import_excel', methods=['POST'])
def import_excel():
    if 'file' not in request.files:
        return jsonify({"error": "لم يتم إرسال أي ملف"}), 400
        
    file = request.files['file']
    days = json.loads(request.form.get('days', '[]'))
    slots = json.loads(request.form.get('slots', '[]'))
    
    if not days or not slots:
        tenant_id = session.get('tenant_id')
        struct_setting = Setting.query.filter_by(key='schedule_structure', tenant_id=tenant_id).first()
        if struct_setting and struct_setting.value:
            structure_data = json.loads(struct_setting.value)
            days = [d['name'] for d in structure_data]
            if structure_data and structure_data[0].get('slots'):
                slots = [f"{s['start']}-{s['end']}" for s in structure_data[0]['slots']]
    
    if not days or not slots:
        return jsonify({"error": "لم يتم العثور على هيكل الأيام والحصص. يرجى إعداده في المرحلة 4 أولاً."}), 400

    try:
        wb = openpyxl.load_workbook(file)
        new_schedule = {}
        new_prof_schedules = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            level_name = str(ws['A1'].value) if ws['A1'].value else sheet_name
            grid = [[ [] for _ in slots ] for _ in days]
            
            for slot_idx, slot_name in enumerate(slots):
                row_idx = slot_idx + 3
                for day_idx, day_name in enumerate(days):
                    col_idx = day_idx + 2 
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if not cell_val or str(cell_val).strip() == "-": continue
                        
                    blocks = str(cell_val).split('===')
                    for block in blocks:
                        lines = [line.strip() for line in block.strip().split('\n') if line.strip()]
                        lec = {"name": "", "teacher_name": "", "room": "", "level": level_name}
                        for line in lines:
                            if line.startswith("مادة:"): lec["name"] = line.replace("مادة:", "").strip()
                            elif line.startswith("أستاذ:"): lec["teacher_name"] = line.replace("أستاذ:", "").strip()
                            elif line.startswith("قاعة:"): lec["room"] = line.replace("قاعة:", "").strip()
                            
                        if lec["name"]: 
                            grid[day_idx][slot_idx].append(lec)
                            t_name = lec["teacher_name"]
                            if t_name and t_name != "None":
                                if t_name not in new_prof_schedules:
                                    new_prof_schedules[t_name] = [[ [] for _ in slots ] for _ in days]
                                new_prof_schedules[t_name][day_idx][slot_idx].append(lec)
                                
            new_schedule[level_name] = grid
            
        return jsonify({"success": True, "schedule": new_schedule, "prof_schedules": new_prof_schedules})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": "فشل في قراءة ملف الإكسل. تأكد من عدم تغيير هيكل الملف. التفاصيل: " + str(e)}), 500


# ================= مهام الخلفية (Celery Tasks) =================

@celery_app.task
def background_generation_task(tenant_id, strict_hierarchy, algorithms, algo_settings):
    log_q = RedisLogQueue(tenant_id)
    scheduling_state = SchedulingStateProxy(tenant_id)
    from flask import current_app as app 
    
    try:
        log_q.put("🚀 بدء جلب البيانات من قاعدة البيانات السحابية المعزولة...")
        
        with app.app_context():
            teachers_list = Teacher.query.filter_by(tenant_id=tenant_id).all()
            teacher_map = {t.id: t.name for t in teachers_list}
            teachers = [{"id": t.id, "name": t.name} for t in teachers_list]
            
            rooms_data = [{"id": r.id, "name": r.name, "type": r.type} for r in Room.query.filter_by(tenant_id=tenant_id).all()]
            levels = [lvl.name for lvl in Level.query.filter_by(tenant_id=tenant_id).order_by(Level.name).all()]
            
            courses_raw = []
            for c in Course.query.filter_by(tenant_id=tenant_id).all():
                courses_raw.append({
                    'id': c.id,
                    'name': c.name,
                    'room_type': c.room_type,
                    'course_nature': c.course_nature,
                    'teacher_name': teacher_map.get(c.teacher_id),
                    'level_names': ",".join([l.name for l in c.levels])
                })
                
            struct_setting = Setting.query.filter_by(key='schedule_structure', tenant_id=tenant_id).first()
            cond_setting = Setting.query.filter_by(key='schedule_conditions', tenant_id=tenant_id).first()
            
            structure_data = json.loads(struct_setting.value) if struct_setting and struct_setting.value else []
            conditions_data = json.loads(cond_setting.value) if cond_setting and cond_setting.value else {}

        if not structure_data:
            raise Exception("لم يتم إعداد هيكل الجدول (المرحلة 4).")

        all_lectures = []
        from collections import defaultdict
        lectures_by_teacher_map = defaultdict(list)
        
        for c in courses_raw:
            lec = dict(c)
            lec['levels'] = lec['level_names'].split(',') if lec['level_names'] else []
            if lec['room_type'] in ['عادية', 'قاعة', 'صغيرة']: lec['room_type'] = 'صغيرة'
            elif lec['room_type'] in ['مدرج', 'كبيرة']: lec['room_type'] = 'كبيرة'
            
            all_lectures.append(lec)
            if lec.get('teacher_name'):
                lectures_by_teacher_map[lec['teacher_name']].append(lec)
        
        lectures_by_teacher_map['__all_lectures__'] = all_lectures

        for r in rooms_data:
            if r['type'] in ['عادية', 'قاعة', 'صغيرة']: r['type'] = 'صغيرة'
            elif r['type'] in ['مدرج', 'كبيرة']: r['type'] = 'كبيرة'

        days = [d['name'] for d in structure_data]
        day_to_idx = {d: i for i, d in enumerate(days)}
        slots = []
        if structure_data and structure_data[0].get('slots'):
            slots = [f"{s['start']}-{s['end']}" for s in structure_data[0]['slots']]
        
        rules_grid = [[[] for _ in slots] for _ in days]
        for d_idx, day_obj in enumerate(structure_data):
            for s_idx, slot_obj in enumerate(day_obj.get('slots', [])):
                for constr in slot_obj.get('constraints', []):
                    rule_type = 'ANY_HALL'
                    if constr['room_rule'] == 'regular': rule_type = 'SMALL_HALLS_ONLY'
                    elif constr['room_rule'] == 'specific': rule_type = 'SPECIFIC_LARGE_HALL'
                    elif constr['room_rule'] == 'none': rule_type = 'NO_HALLS_ALLOWED'
                    
                    rules_grid[d_idx][s_idx].append({
                        'rule_type': rule_type,
                        'levels': constr['levels'],
                        'hall_name': constr['specific_halls'][0] if constr['specific_halls'] else None
                    })

        identifiers_by_level = conditions_data.get('identifiers', {})
        teacher_rules = conditions_data.get('teacher_rules', {})
        global_rules = conditions_data.get('global', {})
        weights = conditions_data.get('weights', {})
        spec_teachers = conditions_data.get('special_teachers', {})
        
        teacher_constraints = {}
        special_constraints = {}
        saturday_teachers = []
        last_slot_restrictions = {}
        
        for t_id_str, rule in teacher_rules.items():
            t_name = next((t['name'] for t in teachers if str(t['id']) == t_id_str), None)
            if not t_name: continue
            
            if rule.get('days'):
                teacher_constraints[t_name] = {'allowed_days': {day_to_idx[d] for d in rule['days'] if d in day_to_idx}}
            
            s_const = {}
            limits = rule.get('limits', [])
            
            if 'always_s2_e4' in limits: s_const['always_s2_to_s4'] = True
            if 's2' in limits: s_const['start_d1_s2'] = True
            if 's3' in limits: s_const['start_d1_s3'] = True
            if 'e3' in limits: s_const['end_s3'] = True
            if 'e4' in limits: s_const['end_s4'] = True
            if rule.get('rule') != 'unspecified':
                rules_map = {'group2': 'يومان متتاليان', 'group3': 'ثلاثة أيام متتالية', 'sep2': 'يومان منفصلان', 'sep3': 'ثلاثة ايام منفصلة'}
                s_const['distribution_rule'] = rules_map.get(rule['rule'], 'غير محدد')
            
            special_constraints[t_name] = s_const

        for t_id_str, spec in spec_teachers.items():
            t_name = next((t['name'] for t in teachers if str(t['id']) == t_id_str), None)
            if not t_name: continue
            if spec.get('allow_saturday'): saturday_teachers.append(t_name)
            if spec.get('prevent_last') == '1': last_slot_restrictions[t_name] = 'last_1'
            elif spec.get('prevent_last') == '2': last_slot_restrictions[t_name] = 'last_2'

        distribution_rule_type = 'strict' if global_rules.get('days_interpretation') == 'strict' else 'allowed'
        max_sess = global_rules.get('max_slots')
        max_sessions_per_day = int(max_sess) if max_sess and max_sess.isdigit() else None
        consecutive_large_hall_rule = global_rules.get('consecutive_hall_ban', 'none')
        
        globally_unavailable_slots = set()
        if global_rules.get('rest_tue_pm') and 'الثلاثاء' in day_to_idx and len(slots) >= 2:
            globally_unavailable_slots.update([(day_to_idx['الثلاثاء'], len(slots)-1), (day_to_idx['الثلاثاء'], len(slots)-2)])
            
        if global_rules.get('rest_last_day_pm') and len(days) > 0 and len(slots) >= 1:
            last_day_idx = len(days) - 1 
            num_slots_to_block = int(global_rules.get('rest_last_day_slots', 2))
            for i in range(1, num_slots_to_block + 1):
                if len(slots) - i >= 0:
                    globally_unavailable_slots.add((last_day_idx, len(slots) - i))

        level_specific_large_rooms = {}
        for lvl, r_id in conditions_data.get('level_amphis', {}).items():
            r_name = next((r['name'] for r in rooms_data if str(r['id']) == str(r_id)), None)
            if r_name: level_specific_large_rooms[lvl] = r_name

        specific_small_room_assignments = {}
        for lvl, r_ids in conditions_data.get('level_small_rooms', {}).items():
            if not isinstance(r_ids, list): r_ids = [r_ids] 
            room_names = []
            for r_id in r_ids:
                r_name = next((r['name'] for r in rooms_data if str(r['id']) == str(r_id)), None)
                if r_name: room_names.append(r_name)
            if room_names:
                is_excl = conditions_data.get('level_exclusive_rooms', {}).get(lvl)
                specific_small_room_assignments[lvl] = [f"EXCL_{name}" if is_excl else name for name in room_names]
        
        specific_small_room_assignments['__GLOBAL_EXCLUSIVE__'] = conditions_data.get('global', {}).get('global_exclusive_rooms', False)
        
        pairs_data = conditions_data.get('pairs', {'share':[], 'noshare':[]})
        teacher_pairs = []
        for p in pairs_data.get('share', []):
            t1 = next((t['name'] for t in teachers if str(t['id']) == str(p[0])), None)
            t2 = next((t['name'] for t in teachers if str(t['id']) == str(p[1])), None)
            if t1 and t2: teacher_pairs.append((t1, t2))
            
        non_sharing_teacher_pairs = []
        for p in pairs_data.get('noshare', []):
            t1 = next((t['name'] for t in teachers if str(t['id']) == str(p[0])), None)
            t2 = next((t['name'] for t in teachers if str(t['id']) == str(p[1])), None)
            if t1 and t2: non_sharing_teacher_pairs.append((t1, t2))

        constraint_severities = {
            'distribution': weights.get('distribution', '10'),
            'non_sharing_days': weights.get('no_share', '10'),
            'saturday_work': weights.get('saturday', '10'),
            'last_slot': weights.get('last_slot', '10'),
            'max_sessions': weights.get('max_daily', '10'),
            'teacher_pairs': weights.get('share_pairs', '10'),
            'consecutive_halls': weights.get('consecutive_halls', '10'),
            'prefer_morning': weights.get('morning_pref', '10'),
            'consecutive_lectures': weights.get('consecutive_lectures', '0'),
            'max_consecutive_lectures_limit': weights.get('max_consecutive_lectures_limit', 2),
            'restricted_day': weights.get('restricted_day', 'السبت')
        }
        
        teacher_constraints['__GLOBAL__'] = {
            'restricted_day': constraint_severities['restricted_day'],
            'weight': constraint_severities['saturday_work']
        }

        for k, v in constraint_severities.items():
            if v == 'strict': constraint_severities[k] = 'hard'
            elif v == '20': constraint_severities[k] = 'high'
            elif v == '10': constraint_severities[k] = 'medium'
            elif v == '1': constraint_severities[k] = 'low'
            elif v == '0': constraint_severities[k] = 'disabled'

        prefer_morning_slots = constraint_severities['prefer_morning'] != 'disabled'
        
        log_q.put("✅ تمت قراءة ومعالجة جميع البيانات والقيود بنجاح!")
        time.sleep(0.5)

        tabu_iterations = int(algo_settings.get('tabu_iterations', 1000))
        tabu_tenure = int(algo_settings.get('tabu_tenure', 10))
        tabu_neighborhood = int(algo_settings.get('tabu_neighborhood', 50))
        lns_iterations = int(algo_settings.get('lns_iterations', 500))
        lns_ruin_factor = float(algo_settings.get('lns_ruin_factor', 20)) / 100.0 
        vns_iterations = int(algo_settings.get('vns_iterations', 300))
        vns_k_max = int(algo_settings.get('vns_k_max', 5))
        lns_stagnation = int(algo_settings.get('lns_stagnation_threshold', 15))
        vns_stagnation = int(algo_settings.get('vns_stagnation_threshold', 15))
        tabu_stagnation = int(algo_settings.get('tabu_stagnation_threshold', 15))

        primary_slots = []
        reserve_slots = []
        half_slots = max(1, len(slots) // 2)
        for d_idx in range(len(days)):
            for s_idx in range(len(slots)):
                if s_idx < half_slots:
                    primary_slots.append((d_idx, s_idx))
                else:
                    reserve_slots.append((d_idx, s_idx))

        log_q.put("\n🚀 جاري بناء الجدول المبدئي السريع (الطماعة)...")
        
        current_solution, final_failures = run_greedy_search_for_best_result(
            log_q=log_q, 
            lectures_sorted=all_lectures,
            days=days, slots=slots, rules_grid=rules_grid, rooms_data=rooms_data, 
            teachers=teachers, all_levels=levels,
            teacher_constraints=teacher_constraints, globally_unavailable_slots=globally_unavailable_slots, 
            special_constraints=special_constraints,
            primary_slots=primary_slots, reserve_slots=reserve_slots, identifiers_by_level=identifiers_by_level, 
            prioritize_primary=True,
            saturday_teachers=saturday_teachers, day_to_idx=day_to_idx, level_specific_large_rooms=level_specific_large_rooms,
            specific_small_room_assignments=specific_small_room_assignments, consecutive_large_hall_rule=consecutive_large_hall_rule, 
            prefer_morning_slots=prefer_morning_slots,
            lectures_by_teacher_map=lectures_by_teacher_map, distribution_rule_type=distribution_rule_type, 
            teacher_pairs=teacher_pairs, constraint_severities=constraint_severities, 
            non_sharing_teacher_pairs=non_sharing_teacher_pairs,
            base_initial_schedule=None
        )
        
        log_q.put(f"✅ تم بناء الجدول المبدئي بنجاح! (باقي {len(final_failures)} أخطاء مرنة)")

        if "tabu" in algorithms and not scheduling_state.get('should_stop'):
            log_q.put(f"\n=== 🔍 بدء البحث المحظور (Tabu Search) بتكرارات: {tabu_iterations} وذاكرة: {tabu_tenure} ===")
            current_solution, final_cost, final_failures = run_tabu_search(
                log_q, all_lectures, days, slots, rooms_data, teachers, levels, 
                identifiers_by_level, special_constraints, teacher_constraints, distribution_rule_type, 
                lectures_by_teacher_map, globally_unavailable_slots, saturday_teachers, teacher_pairs, 
                day_to_idx, rules_grid, scheduling_state, last_slot_restrictions, 
                level_specific_large_rooms, specific_small_room_assignments, constraint_severities, 
                mutation_hard_intensity=algo_settings.get('mutation_hard_intensity', 4), 
                mutation_soft_probability=algo_settings.get('mutation_soft_probability', 0.5),
                max_sessions_per_day=max_sessions_per_day, initial_solution=current_solution, 
                max_iterations=tabu_iterations, tabu_tenure=tabu_tenure, neighborhood_size=tabu_neighborhood,
                consecutive_large_hall_rule=consecutive_large_hall_rule, progress_channel=scheduling_state, 
                prefer_morning_slots=prefer_morning_slots, use_strict_hierarchy=strict_hierarchy, non_sharing_teacher_pairs=non_sharing_teacher_pairs,
                tabu_stagnation_threshold=tabu_stagnation
            )
            
        if "lns" in algorithms and not scheduling_state.get('should_stop'):
            log_q.put(f"\n=== 🌪️ بدء البحث الجواري الواسع (LNS) بتكرارات: {lns_iterations} وتخريب: {lns_ruin_factor*100}% ===")
            current_solution, final_cost, final_failures = run_large_neighborhood_search(
                log_q, all_lectures, days, slots, rooms_data, teachers, levels, 
                identifiers_by_level, special_constraints, teacher_constraints, distribution_rule_type, 
                lectures_by_teacher_map, globally_unavailable_slots, saturday_teachers, teacher_pairs, 
                day_to_idx, rules_grid, max_iterations=lns_iterations, ruin_factor=lns_ruin_factor, prioritize_primary=True,
                mutation_hard_intensity=algo_settings.get('mutation_hard_intensity', 4), 
                mutation_soft_probability=algo_settings.get('mutation_soft_probability', 0.5),
                scheduling_state=scheduling_state, last_slot_restrictions=last_slot_restrictions, 
                level_specific_large_rooms=level_specific_large_rooms, specific_small_room_assignments=specific_small_room_assignments, 
                constraint_severities=constraint_severities, initial_solution=current_solution, 
                max_sessions_per_day=max_sessions_per_day, consecutive_large_hall_rule=consecutive_large_hall_rule, 
                progress_channel=scheduling_state, prefer_morning_slots=prefer_morning_slots, 
                use_strict_hierarchy=strict_hierarchy, non_sharing_teacher_pairs=non_sharing_teacher_pairs,
                lns_stagnation_threshold=lns_stagnation
            )
            
        if "vns" in algorithms and not scheduling_state.get('should_stop'):
            log_q.put(f"\n=== 🌊 بدء البحث الجواري المتغير (VNS) بتكرارات: {vns_iterations} وجوار أقصى: {vns_k_max} ===")
            current_solution, final_cost, final_failures = run_variable_neighborhood_search(
                log_q, all_lectures, days, slots, rooms_data, teachers, levels,
                identifiers_by_level, special_constraints, teacher_constraints, distribution_rule_type,
                lectures_by_teacher_map, globally_unavailable_slots, saturday_teachers, teacher_pairs,
                day_to_idx, rules_grid, max_iterations=vns_iterations, k_max=vns_k_max, prioritize_primary=True,
                mutation_hard_intensity=algo_settings.get('mutation_hard_intensity', 4), 
                mutation_soft_probability=algo_settings.get('mutation_soft_probability', 0.5),
                scheduling_state=scheduling_state, last_slot_restrictions=last_slot_restrictions, 
                level_specific_large_rooms=level_specific_large_rooms, specific_small_room_assignments=specific_small_room_assignments, 
                constraint_severities=constraint_severities, algorithm_settings={'vns_local_search_iterations': 10}, 
                initial_solution=current_solution, max_sessions_per_day=max_sessions_per_day, 
                consecutive_large_hall_rule=consecutive_large_hall_rule, progress_channel=scheduling_state, 
                prefer_morning_slots=prefer_morning_slots, use_strict_hierarchy=strict_hierarchy, non_sharing_teacher_pairs=non_sharing_teacher_pairs,
                vns_stagnation_threshold=vns_stagnation
            )

        if scheduling_state.get('should_stop'):
            log_q.put("\n🛑 تم إيقاف عملية التوزيع من قبل المستخدم.")
            return
        else:
            log_q.put("\n✅ تم الانتهاء من جميع الخوارزميات بنجاح!")

        if final_failures:
            log_q.put("\n" + "="*50)
            log_q.put("📊 تقرير الأخطاء المتبقية في الجدول النهائي:")
            log_q.put("="*50)
            
            missing = [f for f in final_failures if f.get('penalty', 0) >= 1000] 
            hard = [f for f in final_failures if 100 <= f.get('penalty', 0) < 1000] 
            soft = [f for f in final_failures if 0 < f.get('penalty', 0) < 100] 
            
            if missing:
                log_q.put(f"❌ المواد غير المجدولة (نقص): {len(missing)}")
                for f in missing[:10]: log_q.put(f"  - {f.get('course_name')} ({f.get('teacher_name')}): {f.get('reason')}")
                if len(missing) > 10: log_q.put("  ... والمزيد")
            
            if hard:
                log_q.put(f"\n🚫 الأخطاء الصارمة (تعارضات قوية): {len(hard)}")
                for f in hard[:10]: log_q.put(f"  - {f.get('course_name')} ({f.get('teacher_name')}): {f.get('reason')}")
                if len(hard) > 10: log_q.put("  ... والمزيد")
                
            if soft:
                log_q.put(f"\n⚠️ الأخطاء المرنة (تفضيلات لم تتحقق): {len(soft)}")
                for f in soft[:10]: log_q.put(f"  - {f.get('course_name')} ({f.get('teacher_name')}): {f.get('reason')}")
                if len(soft) > 10: log_q.put("  ... والمزيد")
                
            log_q.put("="*50 + "\n")
        else:
            log_q.put("\n🎉 الجدول مثالي! لا توجد أي أخطاء متبقية.")

        log_q.put("جاري تجهيز ملفات التصدير (جداول الأساتذة والقاعات)...")
        
        prof_schedules = {t['name']: [[[] for _ in slots] for _ in days] for t in teachers}
        free_rooms = [[[] for _ in slots] for _ in days]
        
        if current_solution:
            for level, grid in current_solution.items():
                for d, day in enumerate(grid):
                    for s, slot in enumerate(day):
                        for lec in slot:
                            t_name = lec.get('teacher_name')
                            if t_name and t_name in prof_schedules:
                                lec_copy = lec.copy()
                                lec_copy['level'] = level
                                prof_schedules[t_name][d][s].append(lec_copy)
            
            for d in range(len(days)):
                for s in range(len(slots)):
                    busy_rooms = set()
                    for level, grid in current_solution.items():
                        for lec in grid[d][s]:
                            if lec.get('room'): busy_rooms.add(lec['room'])
                    
                    for r in rooms_data:
                        if r['name'] not in busy_rooms:
                            free_rooms[d][s].append(r['name'])
                            
        prof_schedules = {p: g for p, g in prof_schedules.items() if any(lec for day in g for slot in day for lec in slot)}

        final_result = {
            "schedule": current_solution if current_solution else {},
            "prof_schedules": prof_schedules, 
            "free_rooms": free_rooms,         
            "days": days,
            "slots": slots,
            "final_failures": final_failures,        
            "total_lectures": len(all_lectures)      
        }
        
        # حفظ النتيجة في قاعدة البيانات للسماح بنشرها للأساتذة لاحقاً
        with app.app_context():
            res_setting = Setting.query.filter_by(key='schedule_result', tenant_id=tenant_id).first()
            if res_setting: res_setting.value = json.dumps(current_solution)
            else: db.session.add(Setting(key='schedule_result', value=json.dumps(current_solution), tenant_id=tenant_id))
            db.session.commit()

        log_q.put(f"DONE{json.dumps(final_result)}")

    except Exception as e:
        log_q.put(f"\n❌ حدث خطأ فادح أثناء التوزيع:\n{str(e)}")
        log_q.put(traceback.format_exc())
    finally:
        log_q.set_running(False)


@celery_app.task
def background_refinement_task(tenant_id, current_schedule, refinement_level, selected_teachers):
    log_q = RedisLogQueue(tenant_id)
    scheduling_state = SchedulingStateProxy(tenant_id)
    from flask import current_app as app

    try:
        log_q.put("\n🚀 بدء عملية ضغط وتحسين جداول الأساتذة (سد الفجوات)...")
        
        with app.app_context():
            teachers_list = Teacher.query.filter_by(tenant_id=tenant_id).all()
            teacher_map = {t.id: t.name for t in teachers_list}
            teachers = [{"id": t.id, "name": t.name} for t in teachers_list]
            
            rooms_data = [{"id": r.id, "name": r.name, "type": r.type} for r in Room.query.filter_by(tenant_id=tenant_id).all()]
            levels = [lvl.name for lvl in Level.query.filter_by(tenant_id=tenant_id).order_by(Level.name).all()]
            
            for r in rooms_data:
                if r['type'] in ['عادية', 'قاعة', 'صغيرة']: r['type'] = 'صغيرة'
                elif r['type'] in ['مدرج', 'كبيرة']: r['type'] = 'كبيرة'
                
            courses_raw = []
            for c in Course.query.filter_by(tenant_id=tenant_id).all():
                courses_raw.append({
                    'id': c.id,
                    'name': c.name,
                    'room_type': c.room_type,
                    'course_nature': c.course_nature,
                    'teacher_name': teacher_map.get(c.teacher_id),
                    'level_names': ",".join([l.name for l in c.levels])
                })
                
            struct_setting = Setting.query.filter_by(key='schedule_structure', tenant_id=tenant_id).first()
            cond_setting = Setting.query.filter_by(key='schedule_conditions', tenant_id=tenant_id).first()
            
            structure_data = json.loads(struct_setting.value) if struct_setting and struct_setting.value else []
            conditions_data = json.loads(cond_setting.value) if cond_setting and cond_setting.value else {}
        
        all_lectures = []
        from collections import defaultdict
        lectures_by_teacher_map = defaultdict(list)
        
        for c in courses_raw:
            lec = dict(c)
            lec['levels'] = lec['level_names'].split(',') if lec['level_names'] else []
            if lec['room_type'] in ['عادية', 'قاعة', 'صغيرة']: lec['room_type'] = 'صغيرة'
            elif lec['room_type'] in ['مدرج', 'كبيرة']: lec['room_type'] = 'كبيرة'
            
            all_lectures.append(lec)
            if lec.get('teacher_name'):
                lectures_by_teacher_map[lec['teacher_name']].append(lec)
        
        lectures_by_teacher_map['__all_lectures__'] = all_lectures

        days = [d['name'] for d in structure_data]
        day_to_idx = {d: i for i, d in enumerate(days)}
        slots = []
        if structure_data and structure_data[0].get('slots'):
            slots = [f"{s['start']}-{s['end']}" for s in structure_data[0]['slots']]
        
        rules_grid = [[[] for _ in slots] for _ in days]
        for d_idx, day_obj in enumerate(structure_data):
            for s_idx, slot_obj in enumerate(day_obj.get('slots', [])):
                for constr in slot_obj.get('constraints', []):
                    rule_type = 'ANY_HALL'
                    if constr['room_rule'] == 'regular': rule_type = 'SMALL_HALLS_ONLY'
                    elif constr['room_rule'] == 'specific': rule_type = 'SPECIFIC_LARGE_HALL'
                    elif constr['room_rule'] == 'none': rule_type = 'NO_HALLS_ALLOWED'
                    rules_grid[d_idx][s_idx].append({
                        'rule_type': rule_type,
                        'levels': constr['levels'],
                        'hall_name': constr['specific_halls'][0] if constr['specific_halls'] else None
                    })
        
        identifiers_by_level = conditions_data.get('identifiers', {})
        teacher_rules = conditions_data.get('teacher_rules', {})
        global_rules = conditions_data.get('global', {})
        weights = conditions_data.get('weights', {})
        spec_teachers = conditions_data.get('special_teachers', {})
        
        teacher_constraints = {}
        special_constraints = {}
        saturday_teachers = []
        last_slot_restrictions = {}
        
        for t_id_str, rule in teacher_rules.items():
            t_name = next((t['name'] for t in teachers if str(t['id']) == t_id_str), None)
            if not t_name: continue
            if rule.get('days'):
                teacher_constraints[t_name] = {'allowed_days': {day_to_idx[d] for d in rule['days'] if d in day_to_idx}}
            s_const = {}
            limits = rule.get('limits', [])
            
            if 'always_s2_e4' in limits: s_const['always_s2_to_s4'] = True
            if 's2' in limits: s_const['start_d1_s2'] = True
            if 's3' in limits: s_const['start_d1_s3'] = True
            if 'e3' in limits: s_const['end_s3'] = True
            if 'e4' in limits: s_const['end_s4'] = True
            rules_map = {'group2': 'يومان متتاليان', 'group3': 'ثلاثة أيام متتالية', 'sep2': 'يومان منفصلان', 'sep3': 'ثلاثة ايام منفصلة'}
            if rule.get('rule') != 'unspecified': s_const['distribution_rule'] = rules_map.get(rule.get('rule'), 'غير محدد')
            special_constraints[t_name] = s_const

        for t_id_str, spec in spec_teachers.items():
            t_name = next((t['name'] for t in teachers if str(t['id']) == t_id_str), None)
            if not t_name: continue
            if spec.get('allow_saturday'): saturday_teachers.append(t_name)
            if spec.get('prevent_last') == '1': last_slot_restrictions[t_name] = 'last_1'
            elif spec.get('prevent_last') == '2': last_slot_restrictions[t_name] = 'last_2'

        distribution_rule_type = 'strict' if global_rules.get('days_interpretation') == 'strict' else 'allowed'
        max_sess = global_rules.get('max_slots')
        max_sessions_per_day = int(max_sess) if max_sess and max_sess.isdigit() else None
        consecutive_large_hall_rule = global_rules.get('consecutive_hall_ban', 'none')
        
        globally_unavailable_slots = set()
        if global_rules.get('rest_tue_pm') and 'الثلاثاء' in day_to_idx and len(slots) >= 2:
            globally_unavailable_slots.update([(day_to_idx['الثلاثاء'], len(slots)-1), (day_to_idx['الثلاثاء'], len(slots)-2)])
        if global_rules.get('rest_last_day_pm') and len(days) > 0 and len(slots) >= 1:
            last_day_idx = len(days) - 1 
            num_slots_to_block = int(global_rules.get('rest_last_day_slots', 2))
            for i in range(1, num_slots_to_block + 1):
                if len(slots) - i >= 0:
                    globally_unavailable_slots.add((last_day_idx, len(slots) - i))

        level_specific_large_rooms = {}
        for lvl, r_id in conditions_data.get('level_amphis', {}).items():
            r_name = next((r['name'] for r in rooms_data if str(r['id']) == str(r_id)), None)
            if r_name: level_specific_large_rooms[lvl] = r_name
            
        specific_small_room_assignments = {}
        for lvl, r_ids in conditions_data.get('level_small_rooms', {}).items():
            if not isinstance(r_ids, list): r_ids = [r_ids] 
            room_names = []
            for r_id in r_ids:
                r_name = next((r['name'] for r in rooms_data if str(r['id']) == str(r_id)), None)
                if r_name: room_names.append(r_name)
            if room_names:
                is_excl = conditions_data.get('level_exclusive_rooms', {}).get(lvl)
                specific_small_room_assignments[lvl] = [f"EXCL_{name}" if is_excl else name for name in room_names]
        
        specific_small_room_assignments['__GLOBAL_EXCLUSIVE__'] = conditions_data.get('global', {}).get('global_exclusive_rooms', False)
        
        pairs_data = conditions_data.get('pairs', {'share':[], 'noshare':[]})
        teacher_pairs = []
        for p in pairs_data.get('share', []):
            t1 = next((t['name'] for t in teachers if str(t['id']) == str(p[0])), None)
            t2 = next((t['name'] for t in teachers if str(t['id']) == str(p[1])), None)
            if t1 and t2: teacher_pairs.append((t1, t2))
            
        non_sharing_teacher_pairs = []
        for p in pairs_data.get('noshare', []):
            t1 = next((t['name'] for t in teachers if str(t['id']) == str(p[0])), None)
            t2 = next((t['name'] for t in teachers if str(t['id']) == str(p[1])), None)
            if t1 and t2: non_sharing_teacher_pairs.append((t1, t2))
        
        constraint_severities = {
            'distribution': weights.get('distribution', '10'),
            'non_sharing_days': weights.get('no_share', '10'),
            'saturday_work': weights.get('saturday', '10'),
            'last_slot': weights.get('last_slot', '10'),
            'max_sessions': weights.get('max_daily', '10'),
            'teacher_pairs': weights.get('share_pairs', '10'),
            'consecutive_halls': weights.get('consecutive_halls', '10'),
            'prefer_morning': weights.get('morning_pref', '10'),
            'consecutive_lectures': weights.get('consecutive_lectures', '0'),
            'max_consecutive_lectures_limit': weights.get('max_consecutive_lectures_limit', 2)
        }
        for k, v in constraint_severities.items():
            if v == 'strict': constraint_severities[k] = 'hard'
            elif v == '20': constraint_severities[k] = 'high'
            elif v == '10': constraint_severities[k] = 'medium'
            elif v == '1': constraint_severities[k] = 'low'
            elif v == '0': constraint_severities[k] = 'disabled'

        if constraint_severities.get('prefer_morning') == 'disabled':
            constraint_severities['prefer_morning'] = 'low'

        actual_selected_names = []
        if selected_teachers and len(selected_teachers) > 0:
            for st in selected_teachers:
                matched_name = next((t['name'] for t in teachers if str(t['id']) == str(st)), None)
                if matched_name:
                    actual_selected_names.append(matched_name)
                else:
                    actual_selected_names.append(st)
        else:
            actual_selected_names = [t['name'] for t in teachers]

        # تشغيل دالة التحسين
        refined_schedule, refinement_log = refine_and_compact_schedule(
            initial_schedule=current_schedule, log_q=log_q, 
            selected_teachers=actual_selected_names, 
            all_lectures=all_lectures, days=days, slots=slots, rooms_data=rooms_data, teachers=teachers, all_levels=levels, 
            identifiers_by_level=identifiers_by_level, special_constraints=special_constraints, teacher_constraints=teacher_constraints, distribution_rule_type=distribution_rule_type,
            lectures_by_teacher_map=lectures_by_teacher_map, globally_unavailable_slots=globally_unavailable_slots, saturday_teachers=saturday_teachers, teacher_pairs=teacher_pairs,
            day_to_idx=day_to_idx, rules_grid=rules_grid, last_slot_restrictions=last_slot_restrictions, level_specific_large_rooms=level_specific_large_rooms,
            specific_small_room_assignments=specific_small_room_assignments, constraint_severities=constraint_severities, max_sessions_per_day=max_sessions_per_day, 
            consecutive_large_hall_rule=consecutive_large_hall_rule, prefer_morning_slots=True, non_sharing_teacher_pairs=non_sharing_teacher_pairs, 
            refinement_level=refinement_level
        )

        prof_schedules = {t['name']: [[[] for _ in slots] for _ in days] for t in teachers}
        free_rooms = [[[] for _ in slots] for _ in days]
        
        for level_name, grid in refined_schedule.items():
            for d, day_slots in enumerate(grid):
                for s, slot_lectures in enumerate(day_slots):
                    for lec in slot_lectures:
                        t_name = lec.get('teacher_name')
                        if t_name and t_name in prof_schedules:
                            lec_copy = lec.copy()
                            lec_copy['level'] = level_name
                            prof_schedules[t_name][d][s].append(lec_copy)

        for d in range(len(days)):
            for s in range(len(slots)):
                busy_rooms = set()
                for grid in refined_schedule.values():
                    for lec in grid[d][s]:
                        if lec.get('room'): busy_rooms.add(lec['room'])
                for r in rooms_data:
                    if r['name'] not in busy_rooms:
                        free_rooms[d][s].append(r['name'])
                        
        prof_schedules = {p: g for p, g in prof_schedules.items() if any(lec for day in g for slot in day for lec in slot)}

        final_result = {
            "schedule": refined_schedule,
            "prof_schedules": prof_schedules,
            "free_rooms": free_rooms,
            "days": days,
            "slots": slots,
            "final_failures": [],
            "total_lectures": len(all_lectures)
        }
        
        # حفظ النتيجة المهذبة
        with app.app_context():
            res_setting = Setting.query.filter_by(key='schedule_result', tenant_id=tenant_id).first()
            if res_setting: res_setting.value = json.dumps(refined_schedule)
            else: db.session.add(Setting(key='schedule_result', value=json.dumps(refined_schedule), tenant_id=tenant_id))
            db.session.commit()

        log_q.put(f"DONE{json.dumps(final_result)}")

    except Exception as e:
        log_q.put(f"\n❌ حدث خطأ أثناء التحسين:\n{str(e)}")
        log_q.put(traceback.format_exc())
    finally:
        log_q.set_running(False)